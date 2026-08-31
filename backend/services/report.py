"""
A test run written out as a spreadsheet, with the pictures in it.

The results page answers "did it work" while you are looking at it. Handing the
answer to somebody else, or keeping it beside a batch of parts, needs a file --
and a file of numbers with no pictures is not evidence of anything. The images
here are the annotated ones, boxes already drawn, so a row can be read without
opening anything else.

Two sheets, because two questions get asked of this. The first is one row per
image: what it read, how sure it was, and whether that is worth trusting. The
second is one row per detection, for anyone who wants to sort or filter.
"""

import base64
import io
import re
from datetime import datetime

from services.projects import ProjectError

# Below this a detection is reported as uncertain. Not a hard threshold -- the
# score threshold the run used already decided what to report at all -- but the
# line under which a reading should be checked by a person before it is acted
# on. 0.5 is where a detector's own confidence stops being a majority.
UNCERTAIN_BELOW = 0.5

# The images go in at a fixed height so rows line up; wider frames simply get
# more width. Big enough to see a box on a plate, small enough that fifty rows
# is still a file someone will open.
THUMBNAIL_HEIGHT = 90
MAX_ROWS = 500

_DATA_URL = re.compile(r'^data:image/[a-zA-Z+]+;base64,')


def _decode(data_url):
    """The bytes behind a data: URL, or None if it is not one."""
    if not isinstance(data_url, str) or not _DATA_URL.match(data_url):
        return None
    try:
        return base64.b64decode(_DATA_URL.sub('', data_url), validate=True)
    except (ValueError, TypeError):
        return None


def _describe(result):
    """
    What to say about one image in the notes column.

    Three cases worth distinguishing, because they call for different actions:
    nothing found at all, something found but not confidently, and a clean
    read. Saying "0.42" and leaving the reader to decide what that means is
    what a column of numbers already does.
    """
    detections = result.get('detections') or []
    if not detections:
        return 'Nothing found. Either the object is absent or the model missed it.'

    weak = [d for d in detections
            if float(d.get('score') or 0) < UNCERTAIN_BELOW]
    if not weak:
        return ''

    listed = ', '.join(
        f"{d.get('label_name')} at {float(d.get('score') or 0):.2f}"
        for d in sorted(weak, key=lambda d: d.get('score') or 0)[:4])
    if len(weak) == len(detections):
        return f'Every detection is below {UNCERTAIN_BELOW:.2f}: {listed}. Check by eye.'
    return f'{len(weak)} of {len(detections)} below {UNCERTAIN_BELOW:.2f}: {listed}.'


def build_workbook(results, meta=None):
    """
    An .xlsx of a model test, returned as bytes.

    `results` is what the test endpoint returned: one entry per image with its
    detections and an annotated preview as a data URL.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ProjectError(
            'openpyxl is not installed, so spreadsheets cannot be written. '
            'Run: pip install openpyxl', status=500)

    results = list(results or [])
    if not results:
        raise ProjectError('There are no results to export')
    truncated = len(results) > MAX_ROWS
    results = results[:MAX_ROWS]

    meta = meta or {}
    book = Workbook()

    # ── Sheet one: a row per image ─────────────────────────────────────────
    sheet = book.active
    sheet.title = 'Results'

    header_font = Font(bold=True)
    warn_fill = PatternFill('solid', fgColor='FFF3CD')
    bad_fill = PatternFill('solid', fgColor='F8D7DA')

    row = 1
    sheet.cell(row, 1, 'Model test').font = Font(bold=True, size=14)
    row += 1
    for label, value in (
        ('Model', meta.get('model_name')),
        ('Device', meta.get('device')),
        ('Score threshold', meta.get('score_threshold')),
        ('Images', len(results)),
        ('Exported', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ):
        if value in (None, ''):
            continue
        sheet.cell(row, 1, label).font = header_font
        sheet.cell(row, 2, value)
        row += 1

    row += 1
    headers = ['#', 'Image', 'File', 'Reading', 'Objects',
               'Confidence', 'Lowest', 'Notes']
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row, column, title)
        cell.font = header_font
    header_row = row

    widths = {1: 5, 2: 26, 3: 30, 4: 18, 5: 10, 6: 13, 7: 10, 8: 58}
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width

    for index, result in enumerate(results, start=1):
        row += 1
        detections = result.get('detections') or []
        scores = [float(d.get('score') or 0) for d in detections]
        note = _describe(result)

        sheet.cell(row, 1, index)
        sheet.cell(row, 3, result.get('filename', '')).alignment = Alignment(
            vertical='center', wrap_text=True)
        sheet.cell(row, 4, result.get('reading') or '').alignment = Alignment(
            vertical='center')
        sheet.cell(row, 5, len(detections))

        if scores:
            average = sum(scores) / len(scores)
            confidence = sheet.cell(row, 6, average)
            confidence.number_format = '0%'
            lowest = sheet.cell(row, 7, min(scores))
            lowest.number_format = '0%'

        sheet.cell(row, 8, note).alignment = Alignment(vertical='center',
                                                       wrap_text=True)

        # Colour carries the same message as the note, for scanning.
        if not detections:
            fill = bad_fill
        elif note:
            fill = warn_fill
        else:
            fill = None
        if fill:
            for column in range(1, len(headers) + 1):
                sheet.cell(row, column).fill = fill

        image_bytes = _decode(result.get('annotated_image'))
        if image_bytes:
            try:
                picture = XLImage(io.BytesIO(image_bytes))
                scale = THUMBNAIL_HEIGHT / picture.height if picture.height else 1
                picture.height = THUMBNAIL_HEIGHT
                picture.width = max(1, int(picture.width * scale))
                sheet.row_dimensions[row].height = THUMBNAIL_HEIGHT * 0.78
                sheet.add_image(picture, f'B{row}')
            except Exception:  # noqa: BLE001 - a row without its picture still counts
                sheet.cell(row, 2, '(image could not be embedded)')

    sheet.freeze_panes = sheet.cell(header_row + 1, 1)

    if truncated:
        row += 2
        sheet.cell(row, 1,
                   f'Only the first {MAX_ROWS} images are included.').font = Font(
            italic=True)

    # ── Sheet two: a row per detection ─────────────────────────────────────
    detail = book.create_sheet('Detections')
    columns = ['File', 'Order', 'Line', 'Label', 'Confidence',
               'Uncertain', 'x1', 'y1', 'x2', 'y2']
    for column, title in enumerate(columns, start=1):
        detail.cell(1, column, title).font = header_font
    for column, width in {1: 34, 2: 8, 3: 7, 4: 14, 5: 13, 6: 11}.items():
        detail.column_dimensions[get_column_letter(column)].width = width

    line = 1
    for result in results:
        for order, detection in enumerate(result.get('detections') or [], start=1):
            line += 1
            score = float(detection.get('score') or 0)
            box = detection.get('box') or [None] * 4
            detail.cell(line, 1, result.get('filename', ''))
            detail.cell(line, 2, order)
            detail.cell(line, 3, detection.get('line'))
            detail.cell(line, 4, detection.get('label_name'))
            confidence = detail.cell(line, 5, score)
            confidence.number_format = '0%'
            detail.cell(line, 6, 'yes' if score < UNCERTAIN_BELOW else '')
            for offset, value in enumerate(box):
                detail.cell(line, 7 + offset, value)
            if score < UNCERTAIN_BELOW:
                for column in range(1, len(columns) + 1):
                    detail.cell(line, column).fill = warn_fill
    detail.freeze_panes = detail.cell(2, 1)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
