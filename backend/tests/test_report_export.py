"""
A model test written out as a spreadsheet.

A page answers "did it work" while you are looking at it. Handing that answer
to somebody else, or filing it beside a batch of parts, needs a file — and a
file of numbers with no pictures is not evidence of anything, so the annotated
previews go in the sheet.

What is checked here is that the file is a real workbook someone can open, that
the pictures are actually in it, and that a detection nobody should trust is
marked as one rather than sitting in a column of numbers looking like the rest.

    python backend/tests/test_report_export.py
"""
import base64
import io
import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

REPO = Path(__file__).resolve().parents[2]
TMP = Path(tempfile.mkdtemp(prefix='vt_report_'))
os.environ['PROJECTS_ROOT'] = str(TMP / 'projects')
os.environ['DATABASE_URL'] = 'sqlite:///' + (TMP / 't.db').as_posix()
os.environ['REQUIRE_AUTH'] = '0'
sys.path.insert(0, str(REPO / 'backend'))

import cv2                                              # noqa: E402
import numpy as np                                      # noqa: E402
from app import create_app                              # noqa: E402
from services import report                             # noqa: E402

app = create_app()
c = app.test_client()

fails = []


def check(label, cond, detail=''):
    print(f'  {"PASS " if cond else "FAIL "}{label}' + ('' if cond else f'  -> {detail}'))
    if not cond:
        fails.append(label)


def preview(seed):
    """An annotated preview as the test endpoint returns one: a data URL."""
    rng = np.random.default_rng(seed)
    img = np.full((160, 240, 3), 40, np.uint8)
    img = cv2.add(img, rng.integers(0, 20, (160, 240, 3), dtype=np.uint8))
    cv2.rectangle(img, (40, 40), (120, 120), (60, 220, 90), 2)
    encoded = base64.b64encode(cv2.imencode('.jpg', img)[1].tobytes()).decode()
    return 'data:image/jpeg;base64,' + encoded


RESULTS = [
    # A clean read: every detection confident.
    {'filename': 'good.jpg', 'reading': '250', 'detection_count': 3,
     'annotated_image': preview(1),
     'detections': [
         {'label_name': '2', 'score': 0.94, 'box': [10, 20, 60, 120], 'line': 0, 'position': 0},
         {'label_name': '5', 'score': 0.91, 'box': [70, 20, 120, 120], 'line': 0, 'position': 1},
         {'label_name': '0', 'score': 0.88, 'box': [130, 20, 180, 120], 'line': 0, 'position': 2},
     ]},
    # Found something, but not confidently: the case a note exists for.
    {'filename': 'unsure.jpg', 'reading': '8', 'detection_count': 1,
     'annotated_image': preview(2),
     'detections': [
         {'label_name': '8', 'score': 0.31, 'box': [12, 22, 62, 122], 'line': 0, 'position': 0},
     ]},
    # Nothing at all, which is a different thing again.
    {'filename': 'empty.jpg', 'reading': '', 'detection_count': 0,
     'annotated_image': preview(3), 'detections': []},
]

print('== the workbook is written ==')
data = report.build_workbook(RESULTS, meta={
    'model_name': 'best.pt', 'device': 'cuda', 'score_threshold': 0.25})
check('it produced bytes', isinstance(data, bytes) and len(data) > 5000, len(data))
check('and they are a zip, which is what xlsx is', data[:2] == b'PK', data[:4])

print('\n== it opens as a workbook ==')
from openpyxl import load_workbook                      # noqa: E402
book = load_workbook(io.BytesIO(data))
check('with both sheets', book.sheetnames == ['Results', 'Detections'],
      book.sheetnames)

sheet = book['Results']
rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
header_index = next(i for i, row in enumerate(rows) if row[0] == '#')
header = rows[header_index]
body = [row for row in rows[header_index + 1:] if row[0] is not None]

check('the columns are named',
      header[:8] == ['#', 'Image', 'File', 'Reading', 'Objects',
                     'Confidence', 'Lowest', 'Notes'], header[:8])
check('one row per image', len(body) == 3, len(body))

print('\n== what each row says ==')
good, unsure, empty = body
check('the reading is carried through', good[3] == '250', good[3])
check('the object count too', good[4] == 3, good[4])
check('confidence is the average, as a fraction for percent formatting',
      abs(good[5] - (0.94 + 0.91 + 0.88) / 3) < 1e-9, good[5])
check('and the lowest is kept separately', abs(good[6] - 0.88) < 1e-9, good[6])
check('a confident row needs no note', not good[7], good[7])

check('an uncertain detection is called out', 'below' in (unsure[7] or ''),
      unsure[7])
check('naming it and its score',
      '8' in (unsure[7] or '') and '0.31' in (unsure[7] or ''), unsure[7])

check('an image with nothing found says so',
      'Nothing found' in (empty[7] or ''), empty[7])
check('and is distinguished from a weak read', empty[7] != unsure[7])

print('\n== the numbers are formatted as percentages ==')
confidence_cell = sheet.cell(header_index + 2, 6)
check('confidence shows as a percent', confidence_cell.number_format == '0%',
      confidence_cell.number_format)

print('\n== rows are coloured so they can be scanned ==')
clean_fill = sheet.cell(header_index + 2, 1).fill.fgColor.rgb
weak_fill = sheet.cell(header_index + 3, 1).fill.fgColor.rgb
none_fill = sheet.cell(header_index + 4, 1).fill.fgColor.rgb
check('the three cases are not all the same colour',
      len({clean_fill, weak_fill, none_fill}) == 3,
      (clean_fill, weak_fill, none_fill))

print('\n== the pictures are in the file ==')
check('one image per row', len(sheet._images) == 3, len(sheet._images))

print('\n== the second sheet is one row per detection ==')
detail = book['Detections']
detail_rows = [[cell.value for cell in row] for row in detail.iter_rows()]
check('its columns are named',
      detail_rows[0][:6] == ['File', 'Order', 'Line', 'Label',
                             'Confidence', 'Uncertain'], detail_rows[0][:6])
check('four detections across the three images',
      len([r for r in detail_rows[1:] if r[0]]) == 4,
      len([r for r in detail_rows[1:] if r[0]]))
check('the weak one is flagged',
      any(r[5] == 'yes' for r in detail_rows[1:]),
      [r[5] for r in detail_rows[1:]])
check('and the confident ones are not',
      sum(1 for r in detail_rows[1:] if r[5] == 'yes') == 1,
      [r[5] for r in detail_rows[1:]])
check('boxes come through',
      detail_rows[1][6:10] == [10, 20, 60, 120], detail_rows[1][6:10])

print('\n== through the endpoint ==')
r = c.post('/api/models/test/export',
           json={'results': RESULTS, 'model_name': 'best.pt',
                 'device': 'cuda', 'score_threshold': 0.25})
check('the request succeeded', r.status_code == 200, r.get_json())
check('it is offered as a file',
      'attachment' in r.headers.get('Content-Disposition', ''),
      r.headers.get('Content-Disposition'))
check('named after the model and dated',
      '.xlsx' in r.headers.get('Content-Disposition', '')
      and 'best_test_' in r.headers.get('Content-Disposition', ''),
      r.headers.get('Content-Disposition'))
check('with the spreadsheet content type',
      'spreadsheetml' in r.mimetype, r.mimetype)
check('and the same bytes open as a workbook',
      load_workbook(io.BytesIO(r.get_data())).sheetnames == ['Results', 'Detections'])

print('\n== nothing to export ==')
r = c.post('/api/models/test/export', json={'results': []})
check('an empty list is refused', r.status_code == 400, r.status_code)
r = c.post('/api/models/test/export', json={})
check('so is a request with no results at all', r.status_code == 400, r.status_code)

print('\n== a result whose preview is missing or broken ==')
broken = [{'filename': 'x.jpg', 'reading': '1', 'detections': [
    {'label_name': '1', 'score': 0.8, 'box': [1, 2, 3, 4]}],
    'annotated_image': 'not-a-data-url'}]
data = report.build_workbook(broken)
book = load_workbook(io.BytesIO(data))
check('the row is still written', book['Results'].max_row > 1)
check('just without a picture', len(book['Results']._images) == 0,
      len(book['Results']._images))

print('\n' + ('REPORT EXPORT OK' if not fails else f'{len(fails)} FAILED: {fails}'))
if fails:
    print(f'(kept for inspection: {TMP})')
else:
    shutil.rmtree(TMP, ignore_errors=True)
sys.exit(1 if fails else 0)
